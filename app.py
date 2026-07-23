"""
Tim 学习助手 - Flask 应用入口
多学科错题管理、智能复习、统计分析
"""
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, make_response, session
from datetime import datetime, date, timedelta
import os
import io
import threading
import uuid as uuid_mod
import time
from functools import wraps
from flask import g

from config import Config
from database import init_db, query_db, execute_db, rows_to_dicts
from utils import (
    get_or_create_uuid, set_uuid_cookie, ensure_user_config, get_user_config,
    save_image, get_image_data, delete_image_files, allowed_file,
    calculate_next_review, generate_pie_chart, generate_bar_chart,
    generate_line_chart, generate_radar_chart, strip_latex
)
from ocr_providers import get_ocr_provider
from ocr_parser import create_ocr_parser
from doc_import import (
    docx_to_lines, pdf_to_lines_or_images, lines_to_questions, check_doc_ext
)
import shutil
import glob as glob_mod
from api_auth import api_token_required, create_token_record, get_token_prefix
from api_ratelimit import limiter

app = Flask(__name__)
app.config.from_object(Config)
app.permanent_session_lifetime = __import__('datetime').timedelta(hours=24)

# 确保上传目录存在
os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)

# ==================== 安全加固 ====================

# 登录速率限制 {ip: (fail_count, lock_until)}
_login_attempts = {}

@app.after_request
def add_security_headers(resp):
    resp.headers['X-Frame-Options'] = 'DENY'
    resp.headers['X-Content-Type-Options'] = 'nosniff'
    resp.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    return resp


@app.before_request
def csrf_check():
    if 'csrf_token' not in session:
        session['csrf_token'] = str(uuid_mod.uuid4())
    if request.method in ('GET', 'HEAD', 'OPTIONS'):
        return
    if request.endpoint in ('login', 'static') or \
       request.path.startswith('/api/v1/'):
        return
    token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
    if not token or token != session.get('csrf_token'):
        return jsonify({'success': False, 'message': 'CSRF 验证失败'}), 403


# ==================== 登录认证 ====================

def login_required(f):
    @wraps(f)
    def decorated(*a, **kw):
        if 'user_id' not in session:
            return redirect(url_for('login', next=request.path))
        return f(*a, **kw)
    return decorated


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # 速率限制
        ip = request.remote_addr
        now = time.time()
        fails, lock_until = _login_attempts.get(ip, (0, 0))
        if now < lock_until:
            flash(f'登录尝试过多，请 {int(lock_until - now)} 秒后重试', 'danger')
            return render_template('login.html')
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        from werkzeug.security import check_password_hash
        from database import query_db
        user = query_db('SELECT * FROM users WHERE username=?', (username,), one=True)
        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            flash('登录成功', 'success')
            nxt = request.args.get('next', '/')
            return redirect(nxt)
        fails = _login_attempts.get(ip, (0, 0))[0] + 1
        _login_attempts[ip] = (fails, now + 60 if fails >= 3 else 0)
        flash('用户名或密码错误', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('已退出', 'info')
    return redirect(url_for('login'))


@app.errorhandler(500)
def internal_error(e):
    return render_template('error.html', message='服务器内部错误，请稍后重试'), 500


@app.before_request
def before_request():
    """初始化请求上下文"""
    g.user_uuid = get_or_create_uuid()
    g.username = session.get('username', '')

def get_knowledge_points():
    """获取所有知识点列表（去重）"""
    return query_db(
        "SELECT DISTINCT zhishidian FROM mistake_records WHERE zhishidian != '' ORDER BY zhishidian", ())


@app.route('/api/knowledge-points/by-subject')
@login_required
def api_knowledge_points_by_subject():
    """按学科获取知识点列表"""
    xueke = request.args.get('xueke', '').strip()
    if xueke:
        rows = query_db(
            "SELECT DISTINCT zhishidian FROM mistake_records WHERE xueke=? AND zhishidian != '' ORDER BY zhishidian",
            (xueke,))
    else:
        rows = query_db(
            "SELECT DISTINCT zhishidian FROM mistake_records WHERE zhishidian != '' ORDER BY zhishidian", ())
    return jsonify([r['zhishidian'] for r in rows])


def auto_backup():
    """每天首次调用时备份 SQLite 数据库，保留最近 7 天"""
    import time as _time
    backup_dir = '/workspace/backups'
    os.makedirs(backup_dir, exist_ok=True)
    today = date.today().isoformat()
    backup_path = os.path.join(backup_dir, f'study_assistant_{today}.db')
    # 今天已备份则跳过
    if os.path.exists(backup_path):
        return
    try:
        shutil.copy2(Config.DATABASE_PATH, backup_path)
        # 清理超过 7 天的备份
        cutoff = _time.time() - 7 * Config.SECONDS_PER_DAY
        for f in glob_mod.glob(os.path.join(backup_dir, 'study_assistant_*.db')):
            if os.path.getmtime(f) < cutoff:
                os.remove(f)
    except Exception:
        pass


# ==================== 首页仪表盘 ====================

@app.route('/',)
@login_required
def index():
    ensure_user_config(g.user_uuid)
    auto_backup()  # 每天自动备份数据库

    # 统计卡片数据
    total_questions = query_db(
        'SELECT COUNT(*) as cnt FROM mistake_records WHERE 1=1 AND status != ?',
        ('archived',), one=True
    )['cnt']

    active_questions = query_db(
        'SELECT COUNT(*) as cnt FROM mistake_records WHERE 1=1 AND status = ?',
        ('active',), one=True
    )['cnt']

    mastered_questions = query_db(
        'SELECT COUNT(*) as cnt FROM mistake_records WHERE 1=1 AND status = ?',
        ('mastered',), one=True
    )['cnt']

    today_str = date.today().isoformat()
    due_today = query_db(
        '''SELECT COUNT(*) as cnt FROM mistake_records
           WHERE 1=1 AND status = 'active'
           AND next_review_at IS NOT NULL
           AND date(next_review_at) <= ?''',
        (today_str,), one=True
    )['cnt']

    # 本周新增
    week_start = date.today().strftime('%Y-%m-%d 00:00:00',)
    week_new = query_db(
        'SELECT COUNT(*) as cnt FROM mistake_records WHERE 1=1 AND bstudio_create_time >= ?',
        (week_start,), one=True
    )['cnt']

    # 最近错题
    recent_questions = query_db(
        '''SELECT * FROM mistake_records
           WHERE 1=1 AND status = 'active'
           ORDER BY bstudio_create_time DESC LIMIT 5''',
        ()
    )

    # 今日待复习错题
    due_reviews = query_db(
        '''SELECT * FROM mistake_records
           WHERE 1=1 AND status = 'active'
           AND next_review_at IS NOT NULL
           AND date(next_review_at) <= ?
           ORDER BY next_review_at ASC LIMIT 5''',
        (today_str,)
    )

    # 为首页问题生成纯文本摘要
    recent_questions = rows_to_dicts(recent_questions)
    due_reviews = rows_to_dicts(due_reviews)
    for q in recent_questions:
        q['timu_plain'] = strip_latex(q['timu'])
    for q in due_reviews:
        q['timu_plain'] = strip_latex(q['timu'])

    # 学科分布饼图
    xueke_stats = query_db(
        '''SELECT xueke, COUNT(*) as cnt FROM mistake_records
           WHERE 1=1 AND status != 'archived'
           GROUP BY xueke ORDER BY cnt DESC''',
        ()
    )
    pie_chart = None
    if xueke_stats:
        labels = [r['xueke'] for r in xueke_stats]
        values = [r['cnt'] for r in xueke_stats]
        pie_chart = generate_pie_chart(labels, values, '学科错题分布')

    # 复习连击：连续几天有复习记录
    streak = 0
    for i in range(60):
        d = date.today() - __import__('datetime').timedelta(days=i)
        cnt = query_db('SELECT COUNT(*) as c FROM review_logs WHERE review_date=?', (d.isoformat(),), one=True)['c']
        if cnt > 0:
            streak += 1
        else:
            break

    # 今日复习时长
    today_sec = query_db("SELECT COALESCE(SUM(time_spent),0) as s FROM review_logs WHERE review_date=?", (today_str,), one=True)['s']
    today_minutes = round(today_sec / 60, 1)

    # 进行中的学习计划
    active_plans = query_db("SELECT * FROM study_plans WHERE status='in_progress' ORDER BY priority DESC LIMIT 3", ())

    resp = make_response(render_template('index.html',
        total_questions=total_questions,
        active_questions=active_questions,
        mastered_questions=mastered_questions,
        due_today=due_today,
        week_new=week_new,
        recent_questions=recent_questions,
        due_reviews=due_reviews,
        pie_chart=pie_chart,
        streak=streak,
        today_minutes=today_minutes,
        active_plans=active_plans))
    return set_uuid_cookie(resp, g.user_uuid)


# ==================== 粘贴导入 ====================

@app.route('/questions/paste-import', methods=['GET', 'POST'])
@login_required
def paste_import():
    if request.method == 'POST':
        text = request.form.get('text', '')
        xueke = request.form.get('xueke', '数学')
        if not text.strip():
            flash('请粘贴题目内容', 'danger')
            return redirect(url_for('paste_import'))

        # 按题号拆分（匹配 1. 2. 或 (1) (2) 或 一、二、）
        import re
        blocks = re.split(r'(?=^(?:\d+[.、．)）]|[（(]\d+[)）]|[一二三四五六七八九十]+[、.．]))', text.strip(), flags=re.MULTILINE)
        saved = 0
        for block in blocks:
            block = block.strip()
            if not block:
                continue
            # 尝试提取答案
            timu_lines = []
            daan = ''
            for line in block.split('\n'):
                m = re.match(r'^(?:答案|答)[：:]\s*(.*)', line)
                if m:
                    daan = m.group(1).strip()
                else:
                    timu_lines.append(line)
            timu = '\n'.join(timu_lines).strip()
            if timu:
                execute_db(
                    "INSERT INTO mistake_records(uuid,sys_platform,xueke,timu,zhengquedaan,difficulty,status) VALUES(?,'web',?,?,?,3,'active')",
                    (g.user_uuid, xueke, timu, daan))
                saved += 1
        flash(f'成功导入 {saved} 道错题！', 'success')
        return redirect(url_for('question_list'))

    resp = make_response(render_template('paste_import.html', subjects=Config.get_subjects()))
    return set_uuid_cookie(resp, g.user_uuid)


# ==================== 错题 CRUD ====================

@app.route('/questions',)
@login_required
def question_list():

    # 查询参数
    page = request.args.get('page', 1, type=int)
    xueke_filter = request.args.get('xueke', '')
    zhishidian_filter = request.args.get('zhishidian', '')
    status_filter = request.args.get('status', '')
    keyword = request.args.get('keyword', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    # 构建查询（不过滤 uuid，展示全部错题）
    where_clauses = ['1=1']
    params = []

    if xueke_filter:
        where_clauses.append('xueke = ?',)
        params.append(xueke_filter)
    if zhishidian_filter:
        where_clauses.append('zhishidian = ?',)
        params.append(zhishidian_filter)
    if status_filter:
        where_clauses.append('status = ?',)
        params.append(status_filter)
    if keyword:
        where_clauses.append('(timu LIKE ? OR zhishidian LIKE ? OR cuowufenxi LIKE ?)',)
        kw = '%' + keyword + '%'
        params.extend([kw, kw, kw])
    if date_from:
        where_clauses.append("bstudio_create_time >= ? || ' 00:00:00'")
        params.append(date_from)
    if date_to:
        where_clauses.append("bstudio_create_time <= ? || ' 23:59:59'")
        params.append(date_to)

    where_sql = ' AND '.join(where_clauses)

    # 总数
    total = query_db(
        f'SELECT COUNT(*) as cnt FROM mistake_records WHERE {where_sql}',
        params, one=True
    )['cnt']

    # 分页
    offset = (page - 1) * Config.PAGE_SIZE
    questions = query_db(
        f'''SELECT * FROM mistake_records WHERE {where_sql}
            ORDER BY bstudio_create_time DESC
            LIMIT ? OFFSET ?''',
        params + [Config.PAGE_SIZE, offset]
    )

    total_pages = max(1, (total + Config.PAGE_SIZE - 1) // Config.PAGE_SIZE)

    # 为每个问题生成纯文本摘要（去掉 LaTeX 标记）
    questions = rows_to_dicts(questions)
    for q in questions:
        q['timu_plain'] = strip_latex(q['timu'])

    # 获取所有知识点用于过滤下拉框
    knowledge_points = get_knowledge_points()

    resp = make_response(render_template('question_list.html',
        questions=questions,
        page=page,
        total_pages=total_pages,
        total=total,
        xueke_filter=xueke_filter,
        zhishidian_filter=zhishidian_filter,
        status_filter=status_filter,
        keyword=keyword,
        date_from=date_from,
        date_to=date_to,
        knowledge_points=knowledge_points,
        subjects=Config.get_subjects(),
        status_options=Config.STATUS_OPTIONS,
        today_str=date.today().isoformat(),
    ))
    return set_uuid_cookie(resp, g.user_uuid)


@app.route('/questions/add', methods=['GET', 'POST'])
@login_required
def add_question():

    if request.method == 'POST':
        xueke = request.form.get('xueke', '')
        timu = request.form.get('timu', '')
        xueshengdaan = request.form.get('xueshengdaan', '')
        zhengquedaan = request.form.get('zhengquedaan', '')
        cuowufenxi = request.form.get('cuowufenxi', '')
        zhishidian = request.form.get('zhishidian', '')
        difficulty = request.form.get('difficulty', 1, type=int)

        if not xueke or not timu:
            flash('学科和题目内容不能为空', 'danger')
            return redirect(url_for('add_question',))

        mistake_id = execute_db(
            '''INSERT INTO mistake_records
               (uuid, sys_platform, xueke, timu, xueshengdaan, zhengquedaan,
                cuowufenxi, zhishidian, difficulty)
               VALUES (?, 'web', ?, ?, ?, ?, ?, ?, ?)''',
            (g.user_uuid, xueke, timu, xueshengdaan, zhengquedaan,
             cuowufenxi, zhishidian, difficulty)
        )

        # 处理图片上传
        if 'images' in request.files:
            files = request.files.getlist('images',)
            for file in files:
                if file and file.filename and allowed_file(file.filename):
                    save_image(file, mistake_id)

        flash('错题添加成功！', 'success')
        return redirect(url_for('question_detail', question_id=mistake_id))

    knowledge_points = get_knowledge_points()

    resp = make_response(render_template('add_question.html',
        subjects=Config.get_subjects(),
        knowledge_points=knowledge_points
    ))
    return set_uuid_cookie(resp, g.user_uuid)


# ==================== OCR 图片导入 ====================

@app.route('/questions/ocr',)
@login_required
def ocr_import():
    """图片 OCR 导入错题页面"""
    knowledge_points = get_knowledge_points()
    resp = make_response(render_template('ocr_import.html',
        subjects=Config.get_subjects(),
        knowledge_points=knowledge_points
    ))
    return set_uuid_cookie(resp, g.user_uuid)


@app.route('/questions/<int:question_id>',)
@login_required
def question_detail(question_id):
    question = query_db(
        'SELECT * FROM mistake_records WHERE id = ? AND 1=1',
        (question_id,), one=True
    )
    if not question:
        flash('错题不存在', 'danger')
        return redirect(url_for('question_list',))

    # 将 sqlite3.Row 转 dict，并预处理 LaTeX 转义（\$ → $），让 KaTeX 能渲染
    question = dict(question)
    for field in ('timu', 'zhengquedaan', 'xueshengdaan', 'cuowufenxi'):
        if question.get(field):
            question[field] = question[field].replace('\\$', '$')

    # 获取图片
    images = query_db(
        'SELECT * FROM mistake_images WHERE mistake_id = ? ORDER BY id',
        (question_id,)
    )

    # 获取复习记录
    reviews = query_db(
        '''SELECT * FROM review_logs
           WHERE mistake_id = ? AND 1=1
           ORDER BY created_at DESC LIMIT 10''',
        (question_id,)
    )

    # 相似错题推荐：同知识点
    similar = []
    kp = question['zhishidian'] or ''
    if kp:
        similar = query_db(
            'SELECT id, timu, xueke FROM mistake_records WHERE zhishidian=? AND id!=? ORDER BY id DESC LIMIT 5',
            (kp, question_id)
        )
    similar = rows_to_dicts(similar)
    for s in similar:
        s['timu_plain'] = strip_latex(s['timu'])

    # 查询该错题关联的学习计划
    linked_plans = query_db(
        '''SELECT sp.* FROM study_plans sp
           JOIN plan_mistakes pm ON sp.id = pm.plan_id
           WHERE pm.mistake_id = ? AND sp.status NOT IN ('deleted','cancelled')
           ORDER BY sp.priority DESC''',
        (question_id,)
    )

    # 所有活跃计划（用于"添加到计划"下拉）
    all_plans = query_db(
        "SELECT id, title FROM study_plans WHERE status NOT IN ('deleted','cancelled') ORDER BY title",
        ()
    )

    resp = make_response(render_template('question_detail.html',
        question=question,
        images=images,
        reviews=reviews,
        similar=similar,
        linked_plans=linked_plans,
        all_plans=all_plans))
    return set_uuid_cookie(resp, g.user_uuid)


@app.route('/api/questions/<int:question_id>/voice', methods=['GET', 'POST', 'DELETE'])
@login_required
def voice_note(question_id):
    if request.method == 'GET':
        r = query_db('SELECT voice_data FROM mistake_records WHERE id=?', (question_id,), one=True)
        if r and r['voice_data']:
            import base64 as b64
            return send_file(io.BytesIO(b64.b64decode(r['voice_data'])), mimetype='audio/webm')
        return '', 404
    if request.method == 'DELETE':
        execute_db("UPDATE mistake_records SET voice_data='' WHERE id=?", (question_id,))
        return jsonify({'success': True})
    data = request.get_json(silent=True) or {}
    voice_b64 = data.get('voice', '')
    if voice_b64:
        execute_db('UPDATE mistake_records SET voice_data=? WHERE id=?', (voice_b64, question_id))
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': '无音频数据'}), 400


@app.route('/questions/<int:question_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_question(question_id):
    question = query_db(
        'SELECT * FROM mistake_records WHERE id = ? AND 1=1',
        (question_id,), one=True
    )
    if not question:
        flash('错题不存在', 'danger')
        return redirect(url_for('question_list',))

    if request.method == 'POST':
        xueke = request.form.get('xueke', '')
        timu = request.form.get('timu', '')
        xueshengdaan = request.form.get('xueshengdaan', '')
        zhengquedaan = request.form.get('zhengquedaan', '')
        cuowufenxi = request.form.get('cuowufenxi', '')
        zhishidian = request.form.get('zhishidian', '')
        difficulty = request.form.get('difficulty', 1, type=int)
        next_review_at = request.form.get('next_review_at', '') or None

        if not xueke or not timu:
            flash('学科和题目内容不能为空', 'danger')
            return redirect(url_for('edit_question', question_id=question_id))

        execute_db(
            '''UPDATE mistake_records SET
               xueke=?, timu=?, xueshengdaan=?, zhengquedaan=?,
               cuowufenxi=?, zhishidian=?, difficulty=?,
               next_review_at=?,
               updated_at=CURRENT_TIMESTAMP
               WHERE id=?''',
            (xueke, timu, xueshengdaan, zhengquedaan,
             cuowufenxi, zhishidian, difficulty, next_review_at, question_id)
        )

        # 处理图片上传
        if 'images' in request.files:
            files = request.files.getlist('images',)
            for file in files:
                if file and file.filename and allowed_file(file.filename):
                    save_image(file, question_id)

        flash('错题更新成功！', 'success')
        return redirect(url_for('question_detail', question_id=question_id))

    # 获取已有图片
    images = query_db(
        'SELECT * FROM mistake_images WHERE mistake_id = ? ORDER BY id',
        (question_id,)
    )
    knowledge_points = get_knowledge_points()

    resp = make_response(render_template('edit_question.html',
        question=question,
        images=images,
        subjects=Config.get_subjects(),
        knowledge_points=knowledge_points
    ))
    return set_uuid_cookie(resp, g.user_uuid)


def _delete_question_records(question_id):
    """软删除：将错题状态改为 deleted，不物理删除数据。返回是否成功。"""
    # 先确认错题存在且归属当前用户（防越权）
    existing = query_db(
        'SELECT id FROM mistake_records WHERE id = ? AND uuid = ? AND 1=1',
        (question_id, g.user_uuid), one=True)
    if not existing:
        return False
    execute_db(
        "UPDATE mistake_records SET status='deleted', updated_at=CURRENT_TIMESTAMP WHERE id=? AND uuid=?",
        (question_id, g.user_uuid))
    return True


@app.route('/questions/<int:question_id>/delete', methods=['POST'])
@login_required
def delete_question(question_id):
    if not _delete_question_records(question_id):
        return jsonify({'success': False, 'message': '错题不存在'})

    flash('错题已删除', 'success')
    return redirect(url_for('question_list',))


@app.route('/api/questions/batch-delete', methods=['POST'])
@login_required
def batch_delete_questions():
    """批量删除错题：接受 JSON {"ids": [...]}（仅当前用户归属的错题会被删除）。"""
    data = request.get_json(silent=True) or {}
    raw_ids = data.get('ids',)

    if not isinstance(raw_ids, list) or len(raw_ids) == 0:
        return jsonify({'success': False, 'message': '未选择任何错题'}), 400

    # 过滤非 int 并去重，保持顺序
    seen = set()
    ids = []
    for x in raw_ids:
        try:
            qid = int(x)
        except (TypeError, ValueError):
            continue
        if qid not in seen:
            seen.add(qid)
            ids.append(qid)

    if not ids:
        return jsonify({'success': False, 'message': '无效的错题 ID'}), 400

    deleted = 0
    errors = []
    for qid in ids:
        try:
            if _delete_question_records(qid):
                deleted += 1
            else:
                errors.append({'id': qid, 'error': '不存在或无权限'})
        except Exception as e:
            errors.append({'id': qid, 'error': str(e)})

    return jsonify({
        'success': True,
        'deleted': deleted,
        'total': len(ids),
        'errors': errors,
    })


@app.route('/api/questions/batch-update', methods=['POST'])
@login_required
def batch_update_questions():
    """批量更新错题：接受 JSON {ids, action, value}"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids')
    action = data.get('action', '')
    value = data.get('value', '')

    if not isinstance(ids, list) or len(ids) == 0:
        return jsonify({'success': False, 'message': '未选择任何错题'}), 400

    valid_actions = {'status', 'zhishidian', 'next_review_at'}
    if action not in valid_actions:
        return jsonify({'success': False, 'message': f'不支持的批量操作: {action}'}), 400

    if action == 'next_review_at' and not value:
        value = ''

    updated = 0
    for qid in ids:
        try:
            execute_db(
                f'UPDATE mistake_records SET {action}=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
                (value, int(qid))
            )
            updated += 1
        except Exception:
            pass

    return jsonify({'success': True, 'updated': updated, 'total': len(ids)})


# ==================== 回收站（已删除错题管理） ====================

@app.route('/questions/deleted',)
@login_required
def question_deleted():
    page = request.args.get('page', 1, type=int)
    per_page = 20
    offset = (page - 1) * per_page

    # 总数
    total = query_db(
        "SELECT COUNT(*) as cnt FROM mistake_records WHERE status = 'deleted'",
        (), one=True
    )['cnt']

    # 分页查询
    rows = query_db(
        "SELECT * FROM mistake_records WHERE status = 'deleted' ORDER BY bstudio_create_time DESC LIMIT ? OFFSET ?",
        (per_page, offset)
    )
    total_pages = max(1, (total + per_page - 1) // per_page)
    rows = rows_to_dicts(rows)
    for r in rows:
        r['timu_plain'] = strip_latex(r['timu'])
    resp = make_response(render_template('question_deleted.html', rows=rows, page=page, total_pages=total_pages, total=total))
    return set_uuid_cookie(resp, get_or_create_uuid())


@app.route('/api/questions/restore', methods=['POST'])
@login_required
def batch_restore_questions():
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': '未选择'}), 400
    for qid in ids:
        execute_db("UPDATE mistake_records SET status='active', updated_at=CURRENT_TIMESTAMP WHERE id=?", (int(qid),))
    return jsonify({'success': True, 'restored': len(ids)})


@app.route('/api/questions/purge', methods=['POST'])
@login_required
def batch_purge_questions():
    """彻底物理删除已删除的错题"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': '未选择'}), 400
    for qid in ids:
        execute_db('DELETE FROM mistake_images WHERE mistake_id=?', (int(qid),))
        execute_db('DELETE FROM review_logs WHERE mistake_id=?', (int(qid),))
        execute_db('DELETE FROM mistake_records WHERE id=?', (int(qid),))
    return jsonify({'success': True, 'purged': len(ids)})


@app.route('/questions/<int:question_id>/toggle-status', methods=['POST'])
@login_required
def toggle_status(question_id):
    new_status = request.form.get('status', 'active')
    execute_db(
        'UPDATE mistake_records SET status=?, updated_at=CURRENT_TIMESTAMP WHERE id=? AND 1=1',
        (new_status, question_id)
    )
    status_labels = {'active': '活跃', 'archived': '已归档', 'mastered': '已掌握'}
    flash(f'错题状态已更新为「{status_labels.get(new_status, new_status)}」', 'success')
    return redirect(url_for('question_detail', question_id=question_id))


# ==================== 图片 API ====================

@app.route('/api/questions/<int:question_id>/image/<int:image_id>',)
@login_required
def get_image(question_id, image_id):
    image = query_db(
        'SELECT * FROM mistake_images WHERE id = ? AND mistake_id = ?',
        (image_id, question_id), one=True
    )
    if not image:
        return jsonify({'error': '图片不存在'}), 404

    data = get_image_data(image)
    if data is None:
        return jsonify({'error': '图片数据为空'}), 404

    mime = image['mime_type'] or 'image/png'
    return send_file(io.BytesIO(data), mimetype=mime)


@app.route('/api/questions/<int:question_id>/image/<int:image_id>/delete', methods=['POST'])
@login_required
def delete_image(question_id, image_id):
    image = query_db(
        'SELECT * FROM mistake_images WHERE id = ? AND mistake_id = ?',
        (image_id, question_id), one=True
    )
    if image:
        delete_image_files([image])
        execute_db('DELETE FROM mistake_images WHERE id = ?', (image_id,))
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': '图片不存在'})


# ==================== OCR API ====================

# OCR 任务内存存储 {task_id: {status, image_paths, result, error, stage, progress}}
ocr_tasks = {}
ocr_tasks_lock = threading.Lock()


def _run_ocr_task(task_id, image_paths):
    """后台线程执行 OCR 识别任务"""
    try:
        with ocr_tasks_lock:
            ocr_tasks[task_id]['status'] = 'preprocessing'
            ocr_tasks[task_id]['stage'] = 'preprocess'
            ocr_tasks[task_id]['message'] = '正在预处理图片...'

        engine = get_ocr_provider()

        with ocr_tasks_lock:
            ocr_tasks[task_id]['status'] = 'ocr'
            ocr_tasks[task_id]['stage'] = 'ocr'
            ocr_tasks[task_id]['message'] = '正在进行百度智能云 OCR 识别...'

        all_lines = []
        total_images = len(image_paths)
        for idx, path in enumerate(image_paths):
            lines = engine.recognize(path)
            all_lines.extend(lines)
            with ocr_tasks_lock:
                ocr_tasks[task_id]['progress'] = int((idx + 1) / total_images * 100)
                ocr_tasks[task_id]['message'] = f'正在识别第 {idx + 1}/{total_images} 张图片...'

        with ocr_tasks_lock:
            ocr_tasks[task_id]['status'] = 'parsing'
            ocr_tasks[task_id]['stage'] = 'parse'
            ocr_tasks[task_id]['message'] = '正在智能提取题目信息...'

        parser = create_ocr_parser()
        results = parser.parse(all_lines)

        # 构建原始文本用于前端展示
        raw_text = '\n'.join([line['text'] for line in all_lines])

        # 转换为可序列化字典列表
        questions = []
        for r in results:
            questions.append({
                'xueke': r.xueke,
                'timu': r.timu,
                'xueshengdaan': r.xueshengdaan,
                'zhengquedaan': r.zhengquedaan,
                'cuowufenxi': r.cuowufenxi,
                'zhishidian': r.zhishidian,
                'confidence': r.confidence,
                'field_confidences': r.field_confidences,
            })

        with ocr_tasks_lock:
            ocr_tasks[task_id]['status'] = 'done'
            ocr_tasks[task_id]['stage'] = 'done'
            ocr_tasks[task_id]['progress'] = 100
            ocr_tasks[task_id]['message'] = f'识别完成！共 {len(questions)} 道题'
            ocr_tasks[task_id]['questions'] = questions
            ocr_tasks[task_id]['question_count'] = len(questions)
            ocr_tasks[task_id]['raw_ocr_text'] = raw_text

    except Exception as e:
        with ocr_tasks_lock:
            ocr_tasks[task_id]['status'] = 'error'
            ocr_tasks[task_id]['error'] = str(e)
            ocr_tasks[task_id]['message'] = f'识别失败: {str(e)}'


@app.route('/api/ocr/upload', methods=['POST'])
@login_required
def api_ocr_upload():
    """上传图片并启动 OCR 识别"""
    if 'images' not in request.files:
        return jsonify({'success': False, 'error': '未上传图片'}), 400

    files = request.files.getlist('images',)
    if not files or not any(f and f.filename for f in files):
        return jsonify({'success': False, 'error': '未选择文件'}), 400

    # 确保临时目录存在
    os.makedirs(Config.OCR_TEMP_FOLDER, exist_ok=True)

    # 保存图片
    image_paths = []
    task_id = str(uuid_mod.uuid4())

    for file in files:
        if file and file.filename and allowed_file(file.filename):
            ext = os.path.splitext(file.filename)[1] or '.png'
            filename = f"{task_id}_{uuid_mod.uuid4().hex[:8]}{ext}"
            filepath = os.path.join(Config.OCR_TEMP_FOLDER, filename)
            file.save(filepath)
            image_paths.append(filepath)

    if not image_paths:
        return jsonify({'success': False, 'error': '没有有效的图片文件'}), 400

    # 创建 OCR 任务
    with ocr_tasks_lock:
        ocr_tasks[task_id] = {
            'status': 'pending',
            'image_paths': image_paths,
            'result': None,
            'error': None,
            'stage': 'pending',
            'progress': 0,
            'message': '任务已创建，等待处理...',
            'created_at': time.time()
        }

    # 启动后台线程
    thread = threading.Thread(target=_run_ocr_task, args=(task_id, image_paths))
    thread.daemon = True
    thread.start()

    return jsonify({
        'success': True,
        'task_id': task_id,
        'image_count': len(image_paths),
        'message': '图片已上传，正在识别...',
    })


@app.route('/api/ocr/status/<task_id>',)
@login_required
def api_ocr_status(task_id):
    """查询 OCR 任务状态"""
    with ocr_tasks_lock:
        task = ocr_tasks.get(task_id)

    if not task:
        return jsonify({'status': 'error', 'error': '任务不存在或已过期'}), 404

    response = {
        'status': task['status'],
        'stage': task.get('stage', 'pending'),
        'progress': task.get('progress', 0),
        'message': task.get('message', ''),
    }

    if task['status'] == 'done':
        response['questions'] = task.get('questions', [])
        response['question_count'] = task.get('question_count', 0)
        response['raw_ocr_text'] = task.get('raw_ocr_text', '')

    if task['status'] == 'error':
        response['error'] = task.get('error', '未知错误')
        response['raw_ocr_text'] = task.get('raw_ocr_text', '')

    return jsonify(response)


# 定期清理过期任务（每次查询时惰性清理）
def _cleanup_expired_tasks():
    """清理超过有效期的 OCR 任务"""
    now = time.time()
    with ocr_tasks_lock:
        expired = [
            tid for tid, t in ocr_tasks.items()
            if now - t.get('created_at', now) > Config.OCR_TASK_EXPIRE_SECONDS
        ]
        for tid in expired:
            # 清理临时图片
            task = ocr_tasks[tid]
            for path in task.get('image_paths', []):
                try:
                    if os.path.exists(path):
                        os.remove(path)
                except Exception:
                    pass
            del ocr_tasks[tid]


@app.route('/api/ocr/save-batch', methods=['POST'])
@login_required
def api_ocr_save_batch():
    """批量保存 OCR 识别出的多道错题"""

    data = request.get_json()
    if not data or 'questions' not in data:
        return jsonify({'success': False, 'error': '无效的请求数据'}), 400

    questions = data['questions']
    if not isinstance(questions, list) or len(questions) == 0:
        return jsonify({'success': False, 'error': '没有可保存的题目'}), 400

    saved_ids = []
    errors = []

    for idx, q in enumerate(questions):
        xueke = q.get('xueke', '').strip()
        timu = q.get('timu', '').strip()

        if not xueke or not timu:
            errors.append({'index': idx, 'error': '学科或题目内容为空'})
            continue

        try:
            mistake_id = execute_db(
                '''INSERT INTO mistake_records
                   (sys_platform, xueke, timu, xueshengdaan, zhengquedaan,
                    cuowufenxi, zhishidian, difficulty)
                   VALUES (?, 'web', ?, ?, ?, ?, ?, ?, ?)''',
                (xueke, timu,
                    q.get('xueshengdaan', ''),
                    q.get('zhengquedaan', ''),
                    q.get('cuowufenxi', ''),
                    q.get('zhishidian', ''),
                    q.get('difficulty', 3)
                )
            )
            saved_ids.append(mistake_id)
        except Exception as e:
            errors.append({'index': idx, 'error': str(e)})

    return jsonify({
        'success': True,
        'saved_count': len(saved_ids),
        'saved_ids': saved_ids,
        'errors': errors,
    })


# ==================== 文档导入 API（PDF / DOCX） ====================

doc_tasks = {}
doc_tasks_lock = threading.Lock()


def _run_doc_task(task_id, filepath, ext):
    """后台线程：解析文档 → 拆题。支持文本提取与 OCR（扫描 PDF）两种模式。"""
    try:
        with doc_tasks_lock:
            doc_tasks[task_id]['stage'] = 'preprocess'
            doc_tasks[task_id]['message'] = '正在读取文档...'

        lines = []
        image_paths = []
        mode = 'text'

        if ext == 'docx':
            lines = docx_to_lines(filepath)
            with doc_tasks_lock:
                doc_tasks[task_id]['progress'] = 30
                doc_tasks[task_id]['message'] = '已提取文本，正在智能解析...'
        elif ext == 'pdf':
            mode, lines, image_paths = pdf_to_lines_or_images(filepath)

        if mode == 'ocr' and image_paths:
            # 扫描版 PDF：逐页走百度 OCR
            with doc_tasks_lock:
                doc_tasks[task_id]['stage'] = 'ocr'
                doc_tasks[task_id]['message'] = f'扫描件共 {len(image_paths)} 页，正在进行 OCR 识别...'
            provider = get_ocr_provider()
            for i, img_path in enumerate(image_paths):
                page_lines = provider.recognize(img_path)
                lines.extend(page_lines)
                pct = int((i + 1) / len(image_paths) * 80)
                with doc_tasks_lock:
                    doc_tasks[task_id]['progress'] = pct
                    doc_tasks[task_id]['message'] = f'已识别第 {i + 1}/{len(image_paths)} 页...'
        else:
            with doc_tasks_lock:
                doc_tasks[task_id]['progress'] = 50

        with doc_tasks_lock:
            doc_tasks[task_id]['stage'] = 'parse'
            doc_tasks[task_id]['progress'] = 85
            doc_tasks[task_id]['message'] = '正在智能提取题目信息...'

        questions = lines_to_questions(lines)
        raw_text = '\n'.join([l['text'] for l in lines])

        with doc_tasks_lock:
            doc_tasks[task_id]['status'] = 'done'
            doc_tasks[task_id]['stage'] = 'done'
            doc_tasks[task_id]['progress'] = 100
            doc_tasks[task_id]['message'] = f'识别完成，共提取 {len(questions)} 道题目'
            doc_tasks[task_id]['questions'] = questions
            doc_tasks[task_id]['question_count'] = len(questions)
            doc_tasks[task_id]['raw_ocr_text'] = raw_text
    except Exception as e:
        with doc_tasks_lock:
            doc_tasks[task_id]['status'] = 'error'
            doc_tasks[task_id]['message'] = '处理失败'
            doc_tasks[task_id]['error'] = str(e)


@app.route('/questions/doc-import',)
@login_required
def doc_import():
    """文档导入页面"""
    resp = make_response(render_template('doc_import.html',
        subjects=Config.get_subjects()))
    return set_uuid_cookie(resp, g.user_uuid)


@app.route('/api/doc/upload', methods=['POST'])
@login_required
def api_doc_upload():
    """上传 PDF / DOCX 文档并启动解析"""
    if 'doc' not in request.files:
        return jsonify({'success': False, 'error': '未上传文件'}), 400

    file = request.files['doc']
    if not file or not file.filename:
        return jsonify({'success': False, 'error': '未选择文件'}), 400

    ext = check_doc_ext(file.filename)
    if not ext:
        return jsonify({'success': False, 'error': '不支持的文件格式，仅支持 PDF、DOCX'}), 400

    if ext == 'doc':
        return jsonify({
            'success': False,
            'error': '不支持旧版 .doc 格式，请在 Word 中另存为 .docx 后重新导入',
        }), 400

    # 保存文档到临时目录
    os.makedirs(Config.OCR_TEMP_FOLDER, exist_ok=True)
    task_id = str(uuid_mod.uuid4())
    filename = f"{task_id}.{ext}"
    filepath = os.path.join(Config.OCR_TEMP_FOLDER, filename)
    file.save(filepath)

    with doc_tasks_lock:
        doc_tasks[task_id] = {
            'status': 'pending',
            'filepath': filepath,
            'result': None,
            'error': None,
            'stage': 'pending',
            'progress': 0,
            'message': '文档已上传，正在处理...',
            'created_at': time.time()
        }

    thread = threading.Thread(target=_run_doc_task, args=(task_id, filepath, ext))
    thread.daemon = True
    thread.start()

    return jsonify({
        'success': True,
        'task_id': task_id,
        'message': '文档已上传，正在解析...',
    })


@app.route('/api/doc/status/<task_id>',)
@login_required
def api_doc_status(task_id):
    """查询文档导入任务状态（与 OCR 状态同构）"""
    with doc_tasks_lock:
        task = doc_tasks.get(task_id)

    if not task:
        return jsonify({'status': 'error', 'error': '任务不存在或已过期'}), 404

    response = {
        'status': task['status'],
        'stage': task.get('stage', 'pending'),
        'progress': task.get('progress', 0),
        'message': task.get('message', ''),
    }

    if task['status'] == 'done':
        response['questions'] = task.get('questions', [])
        response['question_count'] = task.get('question_count', 0)
        response['raw_ocr_text'] = task.get('raw_ocr_text', '')

    if task['status'] == 'error':
        response['error'] = task.get('error', '未知错误')
        response['raw_ocr_text'] = task.get('raw_ocr_text', '')

    return jsonify(response)

# ==================== 知识点管理 ====================

@app.route('/knowledge-points',)
@login_required
def knowledge_points():

    # 筛选参数
    xueke_filter = request.args.get('xueke', '')
    days_filter = request.args.get('days', '', type=int)
    status_filter = request.args.get('status', '')
    sort = request.args.get('sort', 'total')

    where_clauses = ["zhishidian != ''"]
    params = []

    if xueke_filter:
        where_clauses.append('xueke = ?')
        params.append(xueke_filter)
    if days_filter and days_filter > 0:
        where_clauses.append("bstudio_create_time >= date('now', ? || ' days')")
        params.append(str(-days_filter))
    if status_filter:
        where_clauses.append('status = ?')
        params.append(status_filter)

    where_sql = ' AND '.join(where_clauses)

    order_map = {
        'total': 'total DESC',
        'rate_asc': 'CAST(mastered_count AS REAL) / total ASC',
        'rate_desc': 'CAST(mastered_count AS REAL) / total DESC',
        'name': 'zhishidian ASC',
    }
    order_sql = order_map.get(sort, 'total DESC')

    kps = query_db(
        f'''SELECT zhishidian,
                  COUNT(*) as total,
                  SUM(CASE WHEN status='active' THEN 1 ELSE 0 END) as active_count,
                  SUM(CASE WHEN status='mastered' THEN 1 ELSE 0 END) as mastered_count
           FROM mistake_records
           WHERE {where_sql}
           GROUP BY zhishidian
           ORDER BY {order_sql}''',
        params
    )

    resp = make_response(render_template('knowledge_points.html',
        kps=kps,
        xueke_filter=xueke_filter,
        days_filter=days_filter or '',
        status_filter=status_filter,
        sort=sort,
        subjects=Config.get_subjects(),
        status_options=Config.STATUS_OPTIONS))
    return set_uuid_cookie(resp, g.user_uuid)


@app.route('/api/knowledge-points/add', methods=['POST'])
@login_required
def api_add_knowledge_point():
    """手动添加知识点（写入 base_data + 可选创建一条占位错题记录）"""
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    xueke = (data.get('xueke') or '').strip()
    if not name:
        return jsonify({'success': False, 'message': '知识点名称不能为空'})

    # 写入 base_data（如果不存在）
    existing = query_db(
        "SELECT id FROM base_data WHERE category='knowledge_point' AND name=?",
        (name,), one=True)
    if not existing:
        execute_db(
            "INSERT INTO base_data (category, name, extra, sort_order) VALUES ('knowledge_point', ?, ?, 99)",
            (name, xueke))

    return jsonify({'success': True, 'message': f'知识点「{name}」已添加'})


@app.route('/api/knowledge-points/rename', methods=['POST'])
@login_required
def api_rename_knowledge_point():
    """重命名知识点：批量更新所有关联错题的 zhishidian"""
    data = request.get_json(silent=True) or {}
    old_name = (data.get('old_name') or '').strip()
    new_name = (data.get('new_name') or '').strip()
    if not old_name or not new_name:
        return jsonify({'success': False, 'message': '新旧名称不能为空'})
    if old_name == new_name:
        return jsonify({'success': False, 'message': '新旧名称相同'})

    # 批量更新 mistake_records
    cnt = execute_db(
        "UPDATE mistake_records SET zhishidian=? WHERE zhishidian=?",
        (new_name, old_name))

    # 更新 study_plans
    execute_db(
        "UPDATE study_plans SET zhishidian=? WHERE zhishidian=?",
        (new_name, old_name))

    # 更新 base_data
    execute_db(
        "UPDATE base_data SET name=? WHERE category='knowledge_point' AND name=?",
        (new_name, old_name))

    return jsonify({'success': True, 'message': f'已将「{old_name}」重命名为「{new_name}」，更新 {cnt} 道错题', 'count': cnt})


@app.route('/api/knowledge-points/merge', methods=['POST'])
@login_required
def api_merge_knowledge_points():
    """合并知识点：将源知识点的错题全部归入目标知识点"""
    data = request.get_json(silent=True) or {}
    source = (data.get('source') or '').strip()
    target = (data.get('target') or '').strip()
    if not source or not target:
        return jsonify({'success': False, 'message': '源和目标知识点不能为空'})
    if source == target:
        return jsonify({'success': False, 'message': '不能合并到自身'})

    cnt = execute_db(
        "UPDATE mistake_records SET zhishidian=? WHERE zhishidian=?",
        (target, source))
    execute_db(
        "UPDATE study_plans SET zhishidian=? WHERE zhishidian=?",
        (target, source))
    execute_db(
        "DELETE FROM base_data WHERE category='knowledge_point' AND name=?",
        (source,))

    return jsonify({'success': True, 'message': f'已将 {cnt} 道错题从「{source}」合并到「{target}」', 'count': cnt})


# ==================== 学习计划 ====================

@app.route('/study-plans')
@login_required
def study_plans():
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', '')
    xueke_filter = request.args.get('xueke', '')
    priority_filter = request.args.get('priority', '', type=int)
    per_page = 12

    # 构建查询条件
    where = ["(target_date LIKE ? OR target_date IS NULL OR target_date='')"]
    params = [month + '%']

    if status_filter:
        where.append('status = ?')
        params.append(status_filter)
    if xueke_filter:
        where.append('xueke = ?')
        params.append(xueke_filter)
    if priority_filter:
        where.append('priority >= ?')
        params.append(priority_filter)

    where_sql = ' AND '.join(where)

    # 总数
    total = query_db(
        f"SELECT COUNT(*) as cnt FROM study_plans WHERE {where_sql}",
        params, one=True
    )['cnt']

    offset = (page - 1) * per_page
    total_pages = max(1, (total + per_page - 1) // per_page)

    # 分页查询
    plans = query_db(
        f"SELECT * FROM study_plans WHERE {where_sql} ORDER BY priority DESC, target_date ASC LIMIT ? OFFSET ?",
        params + [per_page, offset]
    )

    # 为每个计划查关联错题数和已掌握数（精确关联 + 模糊兜底）
    plan_mistake_counts = {}
    plan_mastered_counts = {}
    plans = rows_to_dicts(plans)
    for p in plans:
        p['timu_plain'] = ''  # 计划没有 timu 字段
        # 优先从 plan_mistakes 精确表查询
        cnt = query_db(
            "SELECT COUNT(*) as c FROM plan_mistakes WHERE plan_id=?",
            (p['id'],), one=True
        )['c']
        if cnt == 0 and p['zhishidian']:
            # 兜底：模糊匹配（兼容旧数据）
            cnt = query_db(
                "SELECT COUNT(*) as c FROM mistake_records WHERE zhishidian LIKE ?",
                ('%' + p['zhishidian'] + '%',), one=True
            )['c']
        plan_mistake_counts[p['id']] = cnt

        mastered = query_db(
            "SELECT COUNT(*) as c FROM plan_mistakes pm JOIN mistake_records mr ON pm.mistake_id=mr.id WHERE pm.plan_id=? AND mr.status='mastered'",
            (p['id'],), one=True
        )['c']
        if mastered == 0 and p['zhishidian']:
            mastered = query_db(
                "SELECT COUNT(*) as c FROM mistake_records WHERE zhishidian LIKE ? AND status='mastered'",
                ('%' + p['zhishidian'] + '%',), one=True
            )['c']
        plan_mastered_counts[p['id']] = mastered

    # 按日分组构建日历数据
    from collections import defaultdict
    calendar = defaultdict(list)
    for p in plans:
        td = p['target_date'] or ''
        if td:
            calendar[td].append(p)

    # 构建月份导航
    try:
        year, mon = map(int, month.split('-'))
    except ValueError:
        year, mon = date.today().year, date.today().month

    prev_m = f'{year}-{mon-1:02d}' if mon > 1 else f'{year-1}-12'
    next_m = f'{year}-{mon+1:02d}' if mon < 12 else f'{year+1}-01'

    # 当月天数
    import calendar as cal_mod
    days_in_month = cal_mod.monthrange(year, mon)[1]
    first_weekday = cal_mod.monthrange(year, mon)[0]  # 0=周一

    month_days = []
    for i in range(first_weekday):
        month_days.append(None)  # 填充空白
    for d in range(1, days_in_month + 1):
        month_days.append(d)

    resp = make_response(render_template('study_plan.html',
        plans=plans,
        subjects=Config.get_subjects(),
        knowledge_points=get_knowledge_points(),
        calendar=dict(calendar),
        month=month,
        prev_month=prev_m,
        next_month=next_m,
        month_days=month_days,
        month_label=f'{year}年{mon}月',
        today=date.today(),
        today_month=date.today().strftime('%Y-%m'),
        plan_mistake_counts=plan_mistake_counts,
        plan_mastered_counts=plan_mastered_counts,
        page=page,
        total_pages=total_pages,
        total=total,
        status_filter=status_filter,
        xueke_filter=xueke_filter,
        priority_filter=priority_filter))
    return set_uuid_cookie(resp, g.user_uuid)


@app.route('/api/study-plans/list')
@login_required
def api_study_plans_list():
    """AJAX 局部刷新计划列表（日历不动）"""
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', '')
    xueke_filter = request.args.get('xueke', '')
    priority_filter = request.args.get('priority', '', type=int)
    per_page = 12

    where = ["(target_date LIKE ? OR target_date IS NULL OR target_date='')"]
    params = [month + '%']
    if status_filter:
        where.append('status = ?'); params.append(status_filter)
    if xueke_filter:
        where.append('xueke = ?'); params.append(xueke_filter)
    if priority_filter:
        where.append('priority >= ?'); params.append(priority_filter)
    where_sql = ' AND '.join(where)

    total = query_db(f"SELECT COUNT(*) as cnt FROM study_plans WHERE {where_sql}", params, one=True)['cnt']
    offset = (page - 1) * per_page
    total_pages = max(1, (total + per_page - 1) // per_page)

    plans = query_db(
        f"SELECT * FROM study_plans WHERE {where_sql} ORDER BY priority DESC, target_date ASC LIMIT ? OFFSET ?",
        params + [per_page, offset])
    plans = rows_to_dicts(plans)

    plan_mistake_counts = {}
    plan_mastered_counts = {}
    for p in plans:
        cnt = query_db("SELECT COUNT(*) as c FROM plan_mistakes WHERE plan_id=?", (p['id'],), one=True)['c']
        if cnt == 0 and p['zhishidian']:
            cnt = query_db("SELECT COUNT(*) as c FROM mistake_records WHERE zhishidian LIKE ?", ('%'+p['zhishidian']+'%',), one=True)['c']
        plan_mistake_counts[p['id']] = cnt
        mastered = query_db("SELECT COUNT(*) as c FROM plan_mistakes pm JOIN mistake_records mr ON pm.mistake_id=mr.id WHERE pm.plan_id=? AND mr.status='mastered'", (p['id'],), one=True)['c']
        if mastered == 0 and p['zhishidian']:
            mastered = query_db("SELECT COUNT(*) as c FROM mistake_records WHERE zhishidian LIKE ? AND status='mastered'", ('%'+p['zhishidian']+'%',), one=True)['c']
        plan_mastered_counts[p['id']] = mastered

    return render_template('study_plan_list.html',
        plans=plans, plan_mistake_counts=plan_mistake_counts,
        plan_mastered_counts=plan_mastered_counts,
        subjects=Config.get_subjects(),
        page=page, total_pages=total_pages, total=total,
        month=month, status_filter=status_filter,
        xueke_filter=xueke_filter, priority_filter=priority_filter)


@app.route('/study-plans/add', methods=['POST'])
@login_required
def add_study_plan():
    title = request.form.get('title', '')
    description = request.form.get('description', '')
    xueke = request.form.get('xueke', '')
    zhishidian = request.form.get('zhishidian', '')
    target_date = request.form.get('target_date', '')
    priority = request.form.get('priority', 1, type=int)

    if not title:
        flash('计划标题不能为空', 'danger')
        return redirect(url_for('study_plans',))

    execute_db(
        '''INSERT INTO study_plans (uuid, title, description, xueke, zhishidian, target_date, priority)
           VALUES (?, ?, ?, ?, ?, ?, ?)''',
        (g.user_uuid, title, description, xueke, zhishidian, target_date or None, priority)
    )
    flash('学习计划创建成功！', 'success')
    return redirect(url_for('study_plans',))


@app.route('/study-plans/<int:plan_id>/update', methods=['POST'])
@login_required
def update_study_plan(plan_id):
    """更新计划状态或编辑计划内容"""
    # 检查是否仅更新状态（旧方式兼容）
    status = request.form.get('status', '')
    if status and not request.form.get('title'):
        if status == 'completed':
            execute_db(
                'UPDATE study_plans SET status=?, completed_at=CURRENT_TIMESTAMP WHERE id=?',
                (status, plan_id)
            )
        else:
            execute_db(
                'UPDATE study_plans SET status=? WHERE id=?',
                (status, plan_id)
            )
        flash('计划状态已更新', 'success')
        return redirect(url_for('study_plans'))

    # 编辑计划内容
    title = request.form.get('title', '').strip()
    if not title:
        flash('计划标题不能为空', 'danger')
        return redirect(url_for('study_plans'))

    description = request.form.get('description', '').strip()
    xueke = request.form.get('xueke', '').strip()
    zhishidian = request.form.get('zhishidian', '').strip()
    target_date = request.form.get('target_date', '').strip()
    priority = request.form.get('priority', 1, type=int)

    execute_db(
        '''UPDATE study_plans
           SET title=?, description=?, xueke=?, zhishidian=?, target_date=?, priority=?
           WHERE id=?''',
        (title, description, xueke, zhishidian, target_date or None, priority, plan_id)
    )
    flash('计划已更新', 'success')
    return redirect(url_for('study_plans'))


@app.route('/study-plans/<int:plan_id>/delete', methods=['POST'])
@login_required
def delete_study_plan(plan_id):
    execute_db("UPDATE study_plans SET status='deleted' WHERE id=?", (plan_id,))
    flash('计划已删除', 'success')
    return redirect(url_for('study_plans',))


@app.route('/study-plans/<int:plan_id>/pause', methods=['POST'])
@login_required
def pause_study_plan(plan_id):
    execute_db("UPDATE study_plans SET status='paused' WHERE id=?", (plan_id,))
    flash('计划已暂停', 'success')
    return redirect(url_for('study_plans'))


@app.route('/study-plans/<int:plan_id>/resume', methods=['POST'])
@login_required
def resume_study_plan(plan_id):
    execute_db("UPDATE study_plans SET status='in_progress' WHERE id=?", (plan_id,))
    flash('计划已恢复', 'success')
    return redirect(url_for('study_plans'))


@app.route('/study-plans/<int:plan_id>/mistakes')
@login_required
def plan_mistakes(plan_id):
    """管理计划关联的错题"""
    plan = query_db('SELECT * FROM study_plans WHERE id=?', (plan_id,), one=True)
    if not plan:
        flash('计划不存在', 'danger')
        return redirect(url_for('study_plans'))

    # 已关联的错题
    linked = query_db(
        '''SELECT mr.* FROM mistake_records mr
           JOIN plan_mistakes pm ON mr.id = pm.mistake_id
           WHERE pm.plan_id = ? AND 1=1
           ORDER BY mr.bstudio_create_time DESC''',
        (plan_id,)
    )
    linked = rows_to_dicts(linked)
    for q in linked:
        q['timu_plain'] = strip_latex(q['timu'])

    # 推荐错题：优先按知识点，兜底按学科
    suggested = []
    if plan['zhishidian']:
        suggested = query_db(
            '''SELECT mr.* FROM mistake_records mr
               WHERE mr.zhishidian LIKE ? AND 1=1
               AND mr.id NOT IN (SELECT mistake_id FROM plan_mistakes WHERE plan_id=?)
               ORDER BY mr.bstudio_create_time DESC LIMIT 20''',
            ('%' + plan['zhishidian'] + '%', plan_id)
        )
    elif plan['xueke']:
        suggested = query_db(
            '''SELECT mr.* FROM mistake_records mr
               WHERE mr.xueke = ? AND 1=1
               AND mr.id NOT IN (SELECT mistake_id FROM plan_mistakes WHERE plan_id=?)
               ORDER BY mr.bstudio_create_time DESC LIMIT 20''',
            (plan['xueke'], plan_id)
        )
    else:
        suggested = query_db(
            '''SELECT mr.* FROM mistake_records mr
               WHERE 1=1
               AND mr.id NOT IN (SELECT mistake_id FROM plan_mistakes WHERE plan_id=?)
               ORDER BY mr.bstudio_create_time DESC LIMIT 20''',
            (plan_id,)
        )
    suggested = rows_to_dicts(suggested)
    for q in suggested:
        q['timu_plain'] = strip_latex(q['timu'])

    resp = make_response(render_template('plan_mistakes.html',
        plan=plan, linked=linked, suggested=suggested,
        subjects=Config.get_subjects()))
    return set_uuid_cookie(resp, g.user_uuid)


@app.route('/api/study-plans/<int:plan_id>/mistakes/add', methods=['POST'])
@login_required
def api_plan_mistakes_add(plan_id):
    """添加错题到计划"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not isinstance(ids, list) or not ids:
        return jsonify({'success': False, 'message': '未选择错题'})
    added = 0
    for mid in ids:
        try:
            execute_db(
                'INSERT OR IGNORE INTO plan_mistakes (plan_id, mistake_id) VALUES (?, ?)',
                (plan_id, int(mid))
            )
            added += 1
        except Exception:
            pass
    return jsonify({'success': True, 'message': f'已添加 {added} 道错题', 'added': added})


@app.route('/api/study-plans/<int:plan_id>/mistakes/remove', methods=['POST'])
@login_required
def api_plan_mistakes_remove(plan_id):
    """从计划移除错题"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not isinstance(ids, list) or not ids:
        return jsonify({'success': False, 'message': '未选择错题'})
    for mid in ids:
        execute_db(
            'DELETE FROM plan_mistakes WHERE plan_id=? AND mistake_id=?',
            (plan_id, int(mid))
        )
    return jsonify({'success': True, 'message': f'已移除 {len(ids)} 道错题'})


@app.route('/api/study-plans/<int:plan_id>/mistakes/auto', methods=['POST'])
@login_required
def api_plan_mistakes_auto(plan_id):
    """自动将匹配知识点的错题加入计划"""
    plan = query_db('SELECT * FROM study_plans WHERE id=?', (plan_id,), one=True)
    if not plan or not plan['zhishidian']:
        return jsonify({'success': False, 'message': '计划未关联知识点'})
    mistakes = query_db(
        '''SELECT id FROM mistake_records
           WHERE zhishidian LIKE ? AND 1=1
           AND id NOT IN (SELECT mistake_id FROM plan_mistakes WHERE plan_id=?)''',
        ('%' + plan['zhishidian'] + '%', plan_id)
    )
    added = 0
    for m in mistakes:
        try:
            execute_db(
                'INSERT OR IGNORE INTO plan_mistakes (plan_id, mistake_id) VALUES (?, ?)',
                (plan_id, m['id'])
            )
            added += 1
        except Exception:
            pass
    return jsonify({'success': True, 'message': f'自动匹配 {added} 道错题', 'added': added})


@app.route('/api/study-plans/<int:plan_id>/review', methods=['POST'])
@login_required
def api_study_plan_review(plan_id):
    """将学习计划关联的错题加入今日复习（优先精确关联，兜底模糊匹配）"""
    plan = query_db('SELECT * FROM study_plans WHERE id=?', (plan_id,), one=True)
    if not plan:
        return jsonify({'success': False, 'message': '计划不存在'})
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 优先从 plan_mistakes 精确关联
    cnt = execute_db(
        '''UPDATE mistake_records SET next_review_at=?, review_stage=0, status='active'
           WHERE id IN (SELECT mistake_id FROM plan_mistakes WHERE plan_id=?)
           AND status != 'deleted' AND status != 'mastered' ''',
        (now, plan_id))

    # 兜底：知识点模糊匹配
    if cnt == 0 and plan['zhishidian']:
        kp = plan['zhishidian']
        cnt = execute_db(
            "UPDATE mistake_records SET next_review_at=?, review_stage=0, status='active' WHERE zhishidian LIKE ? AND status NOT IN ('deleted','mastered')",
            (now, '%' + kp + '%'))

    return jsonify({'success': True, 'message': f'已加入 {cnt} 道错题到今日复习', 'count': cnt})


# ==================== 复习系统 ====================

@app.route('/review',)
@login_required
def review_plan():
    user_config = get_user_config(g.user_uuid)
    today_str = date.today().isoformat()

    # 筛选 & 排序参数
    xueke_filter = request.args.get('xueke', '')
    sort = request.args.get('sort', 'time')  # 'time' 或 'priority'

    where_clauses = ["status = 'active'", 'next_review_at IS NOT NULL',
                     'date(next_review_at) <= ?']
    params = [today_str]

    if xueke_filter:
        where_clauses.append('xueke = ?')
        params.append(xueke_filter)

    where_sql = ' AND '.join(where_clauses)
    order_sql = 'difficulty DESC, review_stage ASC' if sort == 'priority' else 'next_review_at ASC'

    # 获取今日待复习错题
    due_reviews = query_db(
        f'''SELECT * FROM mistake_records
           WHERE {where_sql}
           ORDER BY {order_sql}
           LIMIT ?''',
        params + [user_config['daily_review_limit']]
    )

    # 为待复习错题加载图片附件
    review_images = {}
    if due_reviews:
        ids = [r['id'] for r in due_reviews]
        placeholders = ','.join(['?'] * len(ids))
        imgs = query_db(
            f'SELECT * FROM mistake_images WHERE mistake_id IN ({placeholders}) ORDER BY id',
            ids
        )
        for img in imgs:
            mid = img['mistake_id']
            if mid not in review_images:
                review_images[mid] = []
            review_images[mid].append(img)

    # 今日已复习数量
    reviewed_today = query_db(
        'SELECT COUNT(*) as cnt FROM review_logs WHERE 1=1 AND review_date = ?',
        (today_str,), one=True
    )['cnt']

    # 复习日历热力图数据（最近 60 天）
    calendar_data = {}
    for i in range(60):
        d = date.today() - __import__('datetime',).timedelta(days=i)
        cnt = query_db(
            'SELECT COUNT(*) as cnt FROM review_logs WHERE 1=1 AND review_date = ?',
            (d.isoformat(),), one=True
        )['cnt']
        if cnt > 0:
            calendar_data[d.isoformat()] = cnt

    resp = make_response(render_template('review_plan.html',
        due_reviews=due_reviews,
        reviewed_today=reviewed_today,
        calendar_data=calendar_data,
        user_config=user_config,
        review_algorithm=user_config['review_algorithm'],
        xueke_filter=xueke_filter,
        sort=sort,
        subjects=Config.get_subjects(),
        review_images=review_images,
    ))
    return set_uuid_cookie(resp, g.user_uuid)


@app.route('/api/review/add-all', methods=['POST'])
@login_required
def api_review_add_all():
    """一键将所有活跃但未排期的错题加入复习（排除已掌握的）"""
    from datetime import datetime
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    execute_db(
        "UPDATE mistake_records SET next_review_at=?, review_stage=0 WHERE status='active' AND next_review_at IS NULL",
        (now,)
    )
    # 也把今天之前的重置到今天（排除已掌握）
    today_str = date.today().isoformat()
    execute_db(
        "UPDATE mistake_records SET next_review_at=? WHERE status='active' AND next_review_at IS NOT NULL AND date(next_review_at) < ?",
        (now, today_str)
    )
    return jsonify({'success': True, 'message': '已将所有未排期及过期错题加入今日复习'})

@app.route('/api/review/today-stats',)
def api_review_today_stats():
    today = date.today().isoformat()
    total = query_db("SELECT COUNT(*) c FROM review_logs WHERE review_date=?",(today,),one=True)['c']
    correct = query_db("SELECT COUNT(*) c FROM review_logs WHERE review_date=? AND result='correct'",(today,),one=True)['c']
    # 连击
    streak = 0
    for i in range(60):
        d = date.today() - __import__('datetime').timedelta(days=i)
        if query_db("SELECT COUNT(*) c FROM review_logs WHERE review_date=?",(d.isoformat(),),one=True)['c'] > 0:
            streak += 1
        else: break
    return jsonify({'total':total, 'correct':correct, 'streak':streak, 'rate':round(correct/total*100,1) if total>0 else 0})


@app.route('/review/config', methods=['GET', 'POST'])
@login_required
def review_config():
    if request.method == 'POST':
        algorithm = request.form.get('algorithm', 'sm2')
        daily_limit = request.form.get('daily_limit', 20, type=int)
        execute_db(
            'UPDATE user_config SET review_algorithm=?, daily_review_limit=? WHERE 1=1',
            (algorithm, daily_limit)
        )
        flash('复习配置已更新', 'success')
        return redirect(url_for('review_plan',))
    return redirect(url_for('review_plan',))


@app.route('/api/review/<int:mistake_id>/submit', methods=['POST'])
@login_required
def submit_review(mistake_id):
    data = request.get_json()
    result = data.get('result', 'correct')
    time_spent = data.get('time_spent', 0)
    notes = data.get('notes', '')

    # 获取当前错题信息
    question = query_db(
        'SELECT * FROM mistake_records WHERE id = ? AND 1=1',
        (mistake_id,), one=True
    )
    if not question:
        return jsonify({'success': False, 'message': '错题不存在'})

    # 计算下次复习日期
    algorithm = question['review_algorithm'] or 'sm2'
    new_stage, next_date = calculate_next_review(
        algorithm, question['review_stage'], result
    )

    # 更新错题复习状态
    execute_db(
        '''UPDATE mistake_records SET
           review_count = review_count + 1,
           last_review_at = CURRENT_TIMESTAMP,
           next_review_at = ?,
           review_stage = ?,
           review_algorithm = ?,
           updated_at = CURRENT_TIMESTAMP
           WHERE id = ? AND 1=1''',
        (next_date.strftime('%Y-%m-%d %H:%M:%S'), new_stage,
         algorithm, mistake_id)
    )

    # 如果连续答对 3 次以上且阶段 >= 4，标记为已掌握
    if result == 'correct' and new_stage >= 4:
        execute_db(
            "UPDATE mistake_records SET status='mastered', updated_at=CURRENT_TIMESTAMP WHERE id=? AND 1=1",
            (mistake_id,)
        )

    # 记录复习日志
    execute_db(
        '''INSERT INTO review_logs (mistake_id, uuid, review_date, result, time_spent, notes)
           VALUES (?, ?, ?, ?, ?, ?)''',
        (mistake_id, g.user_uuid, date.today().isoformat(), result, time_spent, notes)
    )

    return jsonify({
        'success': True,
        'new_stage': new_stage,
        'next_review': next_date.strftime('%Y-%m-%d',),
        'algorithm': algorithm,
        'interval_days': (next_date - datetime.now()).days
    })


# ==================== 历史复习记录 ====================

@app.route('/review/history')
@login_required
def review_history():
    """历史复习记录页面"""
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', '')
    xueke_filter = request.args.get('xueke', '')
    date_range = request.args.get('range', 'all')
    per_page = 20

    # 日期范围
    date_where = ''
    if date_range == '7days':
        date_where = "AND rl.review_date >= date('now','-7 days')"
    elif date_range == '30days':
        date_where = "AND rl.review_date >= date('now','-30 days')"

    # 学科筛选
    xueke_where = ''
    xueke_params = []
    if xueke_filter:
        xueke_where = 'AND mr.xueke = ?'
        xueke_params = [xueke_filter]

    # 结果筛选
    status_where = ''
    status_params = []
    if status_filter in ('correct', 'incorrect', 'partial'):
        status_where = 'AND rl.result = ?'
        status_params = [status_filter]

    base_params = [g.user_uuid] + xueke_params + status_params

    # 总数
    total = query_db(
        f"""SELECT COUNT(*) as cnt FROM review_logs rl
            JOIN mistake_records mr ON rl.mistake_id = mr.id
            WHERE rl.uuid=? {date_where} {xueke_where} {status_where}""",
        base_params, one=True
    )['cnt']
    total_pages = max(1, (total + per_page - 1) // per_page)
    offset = (page - 1) * per_page

    # 分页查询
    reviews = query_db(
        f"""SELECT rl.*, mr.timu, mr.xueke, mr.zhishidian, mr.status as mistake_status
            FROM review_logs rl
            JOIN mistake_records mr ON rl.mistake_id = mr.id
            WHERE rl.uuid=? {date_where} {xueke_where} {status_where}
            ORDER BY rl.created_at DESC
            LIMIT ? OFFSET ?""",
        base_params + [per_page, offset]
    )
    reviews = rows_to_dicts(reviews)
    for r in reviews:
        r['timu_plain'] = strip_latex(r['timu'], max_len=60)

    # 统计摘要
    total_all = query_db(
        "SELECT COUNT(*) as cnt FROM review_logs WHERE uuid=?", (g.user_uuid,), one=True
    )['cnt']
    correct_all = query_db(
        "SELECT COUNT(*) as cnt FROM review_logs WHERE uuid=? AND result='correct'", (g.user_uuid,), one=True
    )['cnt']
    correct_rate = round(correct_all / total_all * 100, 1) if total_all > 0 else 0
    today_cnt = query_db(
        "SELECT COUNT(*) as cnt FROM review_logs WHERE uuid=? AND review_date=date('now')", (g.user_uuid,), one=True
    )['cnt']

    # 连续打卡天数
    streak = 0
    today = date.today()
    for i in range(60):
        check_date = (today - timedelta(days=i)).isoformat()
        cnt = query_db(
            "SELECT COUNT(*) as c FROM review_logs WHERE uuid=? AND review_date=?", (g.user_uuid, check_date), one=True
        )['c']
        if cnt > 0:
            streak += 1
        else:
            if i > 0:  # 今天没复习不算断
                break
            streak = 0

    # 按日期分组
    from collections import defaultdict
    grouped = defaultdict(list)
    for r in reviews:
        grouped[r['review_date']].append(r)

    resp = make_response(render_template('review_history.html',
        reviews=reviews, grouped=dict(grouped),
        subjects=Config.get_subjects(),
        page=page, total_pages=total_pages, total=total,
        status_filter=status_filter, xueke_filter=xueke_filter,
        date_range=date_range,
        total_all=total_all, correct_all=correct_all,
        correct_rate=correct_rate, today_cnt=today_cnt, streak=streak
    ))
    return set_uuid_cookie(resp, g.user_uuid)


@app.route('/api/review/history/list')
@login_required
def api_review_history_list():
    """AJAX 局部刷新复习记录列表"""
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', '')
    xueke_filter = request.args.get('xueke', '')
    date_range = request.args.get('range', 'all')
    per_page = 20

    date_where = ''
    if date_range == '7days':
        date_where = "AND rl.review_date >= date('now','-7 days')"
    elif date_range == '30days':
        date_where = "AND rl.review_date >= date('now','-30 days')"

    xueke_where = ''
    xueke_params = []
    if xueke_filter:
        xueke_where = 'AND mr.xueke = ?'
        xueke_params = [xueke_filter]

    status_where = ''
    status_params = []
    if status_filter in ('correct', 'incorrect', 'partial'):
        status_where = 'AND rl.result = ?'
        status_params = [status_filter]

    base_params = [g.user_uuid] + xueke_params + status_params

    total = query_db(
        f"""SELECT COUNT(*) as cnt FROM review_logs rl
            JOIN mistake_records mr ON rl.mistake_id = mr.id
            WHERE rl.uuid=? {date_where} {xueke_where} {status_where}""",
        base_params, one=True
    )['cnt']
    total_pages = max(1, (total + per_page - 1) // per_page)
    offset = (page - 1) * per_page

    reviews = query_db(
        f"""SELECT rl.*, mr.timu, mr.xueke, mr.zhishidian, mr.status as mistake_status
            FROM review_logs rl
            JOIN mistake_records mr ON rl.mistake_id = mr.id
            WHERE rl.uuid=? {date_where} {xueke_where} {status_where}
            ORDER BY rl.created_at DESC
            LIMIT ? OFFSET ?""",
        base_params + [per_page, offset]
    )
    reviews = rows_to_dicts(reviews)
    for r in reviews:
        r['timu_plain'] = strip_latex(r['timu'], max_len=60)

    from collections import defaultdict
    grouped = defaultdict(list)
    for r in reviews:
        grouped[r['review_date']].append(r)

    return render_template('review_history_list.html',
        reviews=reviews, grouped=dict(grouped),
        page=page, total_pages=total_pages, total=total,
        status_filter=status_filter, xueke_filter=xueke_filter,
        date_range=date_range, subjects=Config.get_subjects())


# ==================== 统计分析 ====================

@app.route('/statistics',)
@login_required
def statistics():

    # 学科分布
    xueke_stats = query_db(
        '''SELECT xueke, COUNT(*) as cnt FROM mistake_records
           WHERE 1=1 AND status != 'archived'
           GROUP BY xueke ORDER BY cnt DESC''',
        ()
    )
    xueke_pie = None
    if xueke_stats:
        xueke_pie = generate_pie_chart(
            [r['xueke'] for r in xueke_stats],
            [r['cnt'] for r in xueke_stats],
            '学科错题分布'
        )

    # 知识点分布（Top 10）
    zp_stats = query_db(
        '''SELECT zhishidian, COUNT(*) as cnt FROM mistake_records
           WHERE 1=1 AND zhishidian != ''
           GROUP BY zhishidian ORDER BY cnt DESC LIMIT 10''',
        ()
    )
    zp_bar = None
    if zp_stats:
        zp_bar = generate_bar_chart(
            [r['zhishidian'][:12] for r in zp_stats],
            [r['cnt'] for r in zp_stats],
            '知识点错题分布 Top 10', '知识点', '错题数'
        )

    # 难度分布雷达图
    difficulty_stats = query_db(
        '''SELECT difficulty, COUNT(*) as cnt FROM mistake_records
           WHERE 1=1 GROUP BY difficulty ORDER BY difficulty''',
        ()
    )
    radar_chart = None
    if difficulty_stats:
        diff_map = {r['difficulty']: r['cnt'] for r in difficulty_stats}
        radar_values = [diff_map.get(i, 0) for i in range(1, 6)]
        radar_chart = generate_radar_chart(
            ['难度1', '难度2', '难度3', '难度4', '难度5'],
            radar_values, '难度分布'
        )

    # 状态统计
    status_stats = query_db(
        '''SELECT status, COUNT(*) as cnt FROM mistake_records
           WHERE 1=1 GROUP BY status''',
        ()
    )
    status_data = {r['status']: r['cnt'] for r in status_stats}
    active_cnt = status_data.get('active', 0)
    archived_cnt = status_data.get('archived', 0)
    mastered_cnt = status_data.get('mastered', 0)
    total = active_cnt + archived_cnt + mastered_cnt

    # 复习趋势（最近 8 周）
    trend_data = []
    trend_labels = []
    for i in range(7, -1, -1):
        week_start = date.today() - __import__('datetime',).timedelta(days=date.today().weekday() + i * 7)
        week_end = week_start + __import__('datetime',).timedelta(days=6)
        cnt = query_db(
            'SELECT COUNT(*) as cnt FROM review_logs WHERE 1=1 AND review_date BETWEEN ? AND ?',
            (week_start.isoformat(), week_end.isoformat()), one=True
        )['cnt']
        trend_labels.append(week_start.strftime('%m/%d',))
        trend_data.append(cnt)

    trend_chart = generate_line_chart(
        trend_labels, trend_data,
        '近 8 周复习趋势', '周', '复习次数'
    ) if any(trend_data) else None

    # 错误类型分布（从 report 合并）
    from error_classifier import classify_batch, ERROR_TYPES
    all_records = query_db("SELECT * FROM mistake_records WHERE 1=1 ORDER BY bstudio_create_time DESC", ())
    error_counts = classify_batch(all_records)
    error_labels = [k for k, v in error_counts.items() if v > 0] or list(error_counts.keys())
    error_values = [error_counts[k] for k in error_labels]
    error_bar = generate_bar_chart(error_labels, error_values, '错误类型分布') if error_labels and any(error_values) else None

    # 掌握程度分布
    status_counter = {'active': 0, 'archived': 0, 'mastered': 0}
    for r in all_records:
        s = r['status'] or 'active'
        if s not in status_counter: status_counter[s] = 0
        status_counter[s] += 1
    ml = [Config.STATUS_OPTIONS.get(k, k) for k in status_counter]
    mv = list(status_counter.values())
    mastery_pie = generate_pie_chart(ml, mv, '掌握程度分布') if total > 0 else None
    mastery_rate = round(status_counter['mastered'] / total * 100, 1) if total else 0

    resp = make_response(render_template('statistics.html',
        xueke_pie=xueke_pie,
        zp_bar=zp_bar,
        radar_chart=radar_chart,
        trend_chart=trend_chart,
        active_cnt=active_cnt,
        archived_cnt=archived_cnt,
        mastered_cnt=mastered_cnt,
        total=total,
        status_stats=status_stats,
        error_labels=error_labels,
        error_values=error_values,
        error_bar=error_bar,
        mastery_labels=ml,
        mastery_values=mv,
        mastery_pie=mastery_pie,
        mastery_rate=mastery_rate,
    ))
    return set_uuid_cookie(resp, g.user_uuid)


# ==================== 错题统计分析报告 ====================

# ==================== 导出 ====================

@app.route('/api/export/excel',)
@login_required
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '错题记录'

        # 表头
        headers = ['ID', '学科', '题目', '学生答案', '标准答案', '错误分析', '知识点',
                   '难度', '复习次数', '状态', '创建时间']
        header_fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 数据
        questions = query_db(
            "SELECT * FROM mistake_records WHERE 1=1 ORDER BY bstudio_create_time DESC",
            ()
        )
        for row_idx, q in enumerate(questions, 2):
            values = [q['id'], q['xueke'], q['timu'], q['xueshengdaan'],
                      q['zhengquedaan'], q['cuowufenxi'], q['zhishidian'],
                      q['difficulty'], q['review_count'],
                      Config.STATUS_OPTIONS.get(q['status'], q['status']),
                      q['bstudio_create_time']]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = thin_border

        # 调整列宽
        col_widths = [6, 8, 40, 25, 25, 30, 15, 8, 8, 10, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'错题记录_{date.today().isoformat()}.xlsx'
        )
    except ImportError:
        flash('导出功能需要 openpyxl 库支持', 'danger')
        return redirect(url_for('question_list',))


# ==================== 错误处理 ====================

@app.errorhandler(404)
def not_found(e):
    return render_template('error.html', code=404,
        message='页面未找到',
        description='您访问的页面不存在，可能已被移动或删除。'), 404


@app.errorhandler(500)
def server_error(e):
    return render_template('error.html', code=500,
        message='服务器错误',
        description='服务器内部发生错误，请稍后重试。'), 500


@app.route('/api/export/pdf',)
@login_required
def export_pdf():
    try:
        from fpdf import FPDF

        questions = query_db(
            "SELECT * FROM mistake_records WHERE 1=1 ORDER BY bstudio_create_time DESC", ()
        )

        pdf = FPDF()
        pdf.add_page()
        font_path = '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'
        pdf.add_font('CJK', '', font_path, uni=True)
        pdf.add_font('CJK', 'B', font_path, uni=True)
        pdf.set_font('CJK', 'B', 14)
        pdf.cell(0, 10, 'Tim 学习助手 - 错题记录', new_x="LMARGIN", new_y="NEXT", align='C')
        pdf.set_font('CJK', '', 8)
        pdf.cell(0, 6, f'导出日期: {date.today().isoformat()}    共 {len(questions)} 道错题', new_x="LMARGIN", new_y="NEXT", align='C')
        pdf.ln(4)

        pdf.set_font('CJK', 'B', 7)
        col_w = [8, 12, 55, 30, 30, 35, 20]
        headers = ['ID', '学科', '题目', '学生答案', '标准答案', '错误分析', '知识点']
        for h, w in zip(headers, col_w):
            pdf.cell(w, 7, h, border=1, align='C')
        pdf.ln()

        pdf.set_font('CJK', '', 6.5)
        for q in questions:
            row = [str(q['id']), q['xueke'] or '',
                   (q['timu'] or '')[:60], (q['xueshengdaan'] or '')[:25],
                   (q['zhengquedaan'] or '')[:25], (q['cuowufenxi'] or '')[:30],
                   (q['zhishidian'] or '')[:18]]
            for val, w in zip(row, col_w):
                pdf.cell(w, 6, val, border=1)
            pdf.ln()

        output = io.BytesIO()
        pdf.output(output)
        output.seek(0)
        return send_file(output, mimetype='application/pdf',
                        as_attachment=True,
                        download_name=f'错题记录_{date.today().isoformat()}.pdf')
    except ImportError:
        flash('PDF 导出需要 fpdf2 库', 'warning')
        return redirect(url_for('question_list',))

@app.route('/api/export/anki',)
@login_required
def export_anki():
    questions = query_db(
        "SELECT * FROM mistake_records WHERE 1=1 ORDER BY bstudio_create_time DESC", ()
    )
    lines = []
    for q in questions:
        front = (q['timu'] or '').replace('\n', '<br>').replace('\t', ' ')
        if q['xueshengdaan']:
            front += '<br><br>我的答案：' + q['xueshengdaan'].replace('\n', '<br>').replace('\t', ' ')
        back = (q['zhengquedaan'] or '').replace('\n', '<br>').replace('\t', ' ')
        if q['cuowufenxi']:
            back += '<br><br>错误分析：' + q['cuowufenxi'].replace('\n', '<br>').replace('\t', ' ')
        if q['zhishidian']:
            back += '<br><br>知识点：' + q['zhishidian']
        lines.append(front + '\t' + back)
    output = io.BytesIO()
    output.write('\n'.join(lines).encode('utf-8'))
    output.seek(0)
    return send_file(output, mimetype='text/plain; charset=utf-8',
                    as_attachment=True,
                    download_name=f'错题_Anki导入_{date.today().isoformat()}.txt')


# ==================== 外部 API v1 ====================

@app.route('/api/v1/ping', methods=['GET'])
@api_token_required
def api_v1_ping():
    """API 连通性测试"""
    remaining = limiter.remaining(
        g.api_token_id,
        rate=g.api_token_rate_limit,
        burst=getattr(Config, 'API_RATE_LIMIT_BURST', 10)
    )
    return jsonify({
        'success': True,
        'data': {
            'message': 'pong',
            'username': g.api_username,
            'timestamp': datetime.now().isoformat(),
        },
        'rate_limit': {
            'remaining': remaining,
            'limit': g.api_token_rate_limit,
        }
    })


@app.route('/api/v1/questions/import', methods=['POST'])
@api_token_required
def api_v1_questions_import():
    """
    批量导入错题
    POST /api/v1/questions/import
    Authorization: Bearer tim_xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx
    Content-Type: application/json

    请求体:
    {
        "questions": [
            {
                "xueke": "数学",           // 必填
                "timu": "题目内容",         // 必填
                "xueshengdaan": "学生答案", // 可选
                "zhengquedaan": "正确答案", // 可选
                "cuowufenxi": "错误分析",   // 可选
                "zhishidian": "知识点",     // 可选
                "difficulty": 3            // 可选，1-5，默认 1
            }
        ]
    }
    """
    # ---- 限速检查 ----
    allowed, retry_after = limiter.check(
        g.api_token_id,
        rate=g.api_token_rate_limit,
        burst=getattr(Config, 'API_RATE_LIMIT_BURST', 10)
    )
    if not allowed:
        return jsonify({
            'success': False,
            'error': 'RATE_LIMITED',
            'message': f'请求过于频繁，请在 {retry_after} 秒后重试',
            'retry_after': retry_after
        }), 429

    # ---- 解析请求 ----
    data = request.get_json(silent=True)
    if not data or 'questions' not in data:
        return jsonify({
            'success': False,
            'error': 'INVALID_REQUEST',
            'message': '请求体必须包含 questions 数组'
        }), 400

    questions = data['questions']
    if not isinstance(questions, list) or len(questions) == 0:
        return jsonify({
            'success': False,
            'error': 'EMPTY_BATCH',
            'message': 'questions 不能为空'
        }), 400

    max_batch = getattr(Config, 'API_IMPORT_MAX_BATCH_SIZE', 200)
    if len(questions) > max_batch:
        return jsonify({
            'success': False,
            'error': 'BATCH_TOO_LARGE',
            'message': f'单次最多导入 {max_batch} 道错题，当前 {len(questions)} 道'
        }), 400

    # ---- 逐条验证和插入 ----
    api_uuid = f'api_user_{g.api_user_id}'
    valid_subjects = Config.get_subjects()

    saved_ids = []
    errors = []

    for idx, q in enumerate(questions):
        xueke = (q.get('xueke') or '').strip()
        timu = (q.get('timu') or '').strip()

        # 必填字段验证
        if not xueke or not timu:
            errors.append({'index': idx, 'error': 'xueke 和 timu 为必填字段'})
            continue

        # 学科合法性验证
        if xueke not in valid_subjects:
            errors.append({
                'index': idx,
                'error': f'无效的学科 "{xueke}"，可选: {", ".join(valid_subjects)}'
            })
            continue

        # difficulty 范围验证
        difficulty = q.get('difficulty', 1)
        try:
            difficulty = int(difficulty)
            if difficulty < 1 or difficulty > 5:
                raise ValueError
        except (TypeError, ValueError):
            errors.append({
                'index': idx,
                'error': f'difficulty 必须为 1-5 的整数，当前值: {difficulty}'
            })
            continue

        try:
            mistake_id = execute_db(
                '''INSERT INTO mistake_records
                   (uuid, sys_platform, xueke, timu, xueshengdaan, zhengquedaan,
                    cuowufenxi, zhishidian, difficulty)
                   VALUES (?, 'api', ?, ?, ?, ?, ?, ?, ?)''',
                (api_uuid, xueke, timu,
                 (q.get('xueshengdaan') or '').strip(),
                 (q.get('zhengquedaan') or '').strip(),
                 (q.get('cuowufenxi') or '').strip(),
                 (q.get('zhishidian') or '').strip(),
                 difficulty)
            )
            saved_ids.append(mistake_id)
        except Exception as e:
            errors.append({'index': idx, 'error': f'数据库错误: {str(e)}'})

    # ---- 构建响应 ----
    remaining = limiter.remaining(
        g.api_token_id,
        rate=g.api_token_rate_limit,
        burst=getattr(Config, 'API_RATE_LIMIT_BURST', 10)
    )

    return jsonify({
        'success': True,
        'data': {
            'total': len(questions),
            'saved_count': len(saved_ids),
            'saved_ids': saved_ids,
            'errors': errors,
        },
        'rate_limit': {
            'remaining': remaining,
            'limit': g.api_token_rate_limit,
        }
    })


# ==================== API Token 管理（Web 页面） ====================

@app.route('/settings/api-tokens')
@login_required
def api_tokens_page():
    """API Token 管理页面"""
    user_id = session['user_id']
    tokens = query_db(
        '''SELECT id, token_prefix, name, description, is_active,
                  rate_limit, last_used_at, created_at, expires_at
           FROM api_tokens WHERE user_id = ? ORDER BY is_active DESC, created_at DESC''',
        (user_id,)
    )
    resp = make_response(render_template('api_tokens.html', tokens=tokens))
    return set_uuid_cookie(resp, g.user_uuid)


@app.route('/api/settings/api-tokens', methods=['POST'])
@login_required
def api_create_token():
    """创建新的 API Token（仅返回一次明文）"""
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    description = (data.get('description') or '').strip()
    rate_limit = int(data.get('rate_limit', 60))

    if rate_limit < 1 or rate_limit > 10000:
        return jsonify({'success': False, 'message': '速率限制范围: 1-10000 次/分钟'}), 400

    user_id = session['user_id']
    full_token, token_id = create_token_record(
        user_id=user_id,
        name=name,
        description=description,
        rate_limit=rate_limit
    )

    return jsonify({
        'success': True,
        'data': {
            'token': full_token,
            'id': token_id,
            'prefix': get_token_prefix(full_token),
            'warning': '请立即复制保存 Token，关闭后无法再次查看明文',
        }
    })


@app.route('/api/settings/api-tokens/<int:token_id>/revoke', methods=['POST'])
@login_required
def api_revoke_token(token_id):
    """撤销（停用）API Token"""
    user_id = session['user_id']
    execute_db(
        'UPDATE api_tokens SET is_active = 0 WHERE id = ? AND user_id = ?',
        (token_id, user_id)
    )
    return jsonify({'success': True, 'message': 'Token 已停用'})


# ==================== 启动 ====================

if __name__ == '__main__':
    print("=" * 50)
    print("  Tim 学习助手 v1.0")
    print("  多学科错题管理 | 智能复习 | 统计分析")
    print("=" * 50)
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=os.environ.get('FLASK_DEBUG', '1') == '1')
